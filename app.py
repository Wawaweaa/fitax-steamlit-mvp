import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# 页面配置
st.set_page_config(
    page_title="电商数据处理系统",
    page_icon="📊",
    layout="wide"
)

# ==================== 配置：平台状态 ====================
PLATFORM_CONFIG = {
    '小红书': {
        'enabled': True,
        'icon': '🔴',
        'status': '已上线',
        'processor': 'process_xiaohongshu'
    },
    '抖音': {
        'enabled': False,
        'icon': '🎵',
        'status': '开发中',
        'processor': 'process_douyin'
    },
    '视频号': {
        'enabled': False,
        'icon': '📹',
        'status': '开发中',
        'processor': 'process_shipinhao'
    }
}

# ==================== 密码保护 ====================
def check_password():
    """简单的密码保护"""
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
    
    if not st.session_state.authenticated:
        st.title("🔐 电商数据处理系统")
        st.markdown("### 请登录以继续")
        
        password = st.text_input("访问密码", type="password", key="password_input")
        
        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            if st.button("登录", use_container_width=True):
                if password == "ecommerce2025":
                    st.session_state.authenticated = True
                    st.rerun()
                else:
                    st.error("❌ 密码错误")
        
        with col2:
            st.markdown("*忘记密码？请联系管理员*")
        
        return False
    
    return True

# ==================== 文件识别 ====================
def identify_files(uploaded_files):
    """识别上传的文件类型"""
    result = {
        'settlement': None,
        'orders': None,
        'settlement_name': None,
        'orders_name': None
    }
    
    for uploaded_file in uploaded_files:
        try:
            uploaded_file.seek(0)
            df = pd.read_excel(uploaded_file, nrows=5)
            columns = set(df.columns)
            
            # 识别结算明细（关键字段：结算时间、佣金总额、商品实付/实退）
            if '结算时间' in columns and '佣金总额' in columns and '商品实付/实退' in columns:
                result['settlement'] = uploaded_file
                result['settlement_name'] = uploaded_file.name
            # 识别订单数据（关键字段：商家编码、商品总价(元)、SKU件数）
            elif '商家编码' in columns and '商品总价(元)' in columns and 'SKU件数' in columns:
                result['orders'] = uploaded_file
                result['orders_name'] = uploaded_file.name
            
            uploaded_file.seek(0)
                
        except Exception as e:
            st.warning(f"⚠️ 无法读取文件 {uploaded_file.name}: {e}")
    
    return result

# ==================== 小红书数据处理（完整版）====================
def process_xiaohongshu_data(settlement_file, orders_file, year, month):
    """处理小红书数据 - 完整版本，包含所有计算和Excel公式"""
    
    # 读取数据
    settlement_file.seek(0)
    orders_file.seek(0)
    
    xhs_settlement = pd.read_excel(settlement_file)
    xhs_orders = pd.read_excel(orders_file)
    
    # 创建订单查找字典
    order_lookup = {}
    for _, row in xhs_orders.iterrows():
        key = f"{row['订单号']}_{row['规格ID']}"
        order_lookup[key] = {
            '商家编码': row.get('商家编码', ''),
            '商品总价(元)': row.get('商品总价(元)', 0),
            'SKU件数': row.get('SKU件数', 1)
        }
    
    # 保持原始顺序
    xhs_dec = xhs_settlement.copy()
    xhs_dec['_original_index'] = range(len(xhs_dec))
    
    # 计算订单行数和订单序位
    xhs_dec['订单行数'] = xhs_dec.groupby('订单号')['订单号'].transform('count')
    xhs_dec['订单序位'] = xhs_dec.groupby('订单号').cumcount() + 1
    
    # 创建查找键
    xhs_dec['lookup_key'] = xhs_dec['订单号'].astype(str) + '_' + xhs_dec['规格ID'].astype(str)
    
    # 查找平台商品编码
    def get_merchant_code(row):
        lookup_key = row['lookup_key']
        if lookup_key in order_lookup:
            return order_lookup[lookup_key]['商家编码']
        return ''
    
    xhs_dec['平台商品编码'] = xhs_dec.apply(get_merchant_code, axis=1)
    
    # 商品编码：提取平台商品编码中"-"之前的部分
    def extract_product_code(code):
        if pd.isna(code) or code == '':
            return ''
        code_str = str(code)
        if '-' in code_str:
            return code_str.split('-')[0]
        return code_str
    
    xhs_dec['商品编码'] = xhs_dec['平台商品编码'].apply(extract_product_code)
    
    # 辅助函数：转换金额
    def to_float(x):
        if pd.isna(x):
            return 0.0
        if isinstance(x, str):
            x = x.replace('¥', '').replace(',', '').strip()
            if x == '':
                return 0.0
        return float(x)
    
    # 计算销售数量
    def calc_sales_qty(row):
        lookup_key = row['lookup_key']
        if lookup_key not in order_lookup:
            return 0
        
        info = order_lookup[lookup_key]
        total_price = to_float(info['商品总价(元)'])
        sku_count = to_float(info['SKU件数'])
        
        if sku_count == 0 or total_price == 0:
            return 0
        
        unit_price = total_price / sku_count
        
        actual_amount = (
            to_float(row.get('商品实付/实退', 0)) +
            to_float(row.get('商家优惠', 0)) +
            to_float(row.get('平台优惠补贴', 0))
        )
        
        if unit_price == 0:
            return 0
        
        ratio = actual_amount / unit_price
        
        if abs(ratio) < 0.15:
            return 0
        else:
            if ratio > 0:
                return int(np.ceil(ratio))
            else:
                return int(np.floor(ratio))
    
    xhs_dec['销售数量'] = xhs_dec.apply(calc_sales_qty, axis=1)
    
    # 计算应收客户
    xhs_dec['应收客户'] = xhs_dec['商品实付/实退'].apply(to_float)
    
    # 计算应收平台
    xhs_dec['应收平台'] = xhs_dec['平台优惠补贴'].apply(to_float)
    
    # 计算收：价外收费
    xhs_dec['运费_数值'] = xhs_dec['运费'].apply(to_float)
    
    # 计算每个订单中销售数量>0的行数
    positive_counts = xhs_dec[xhs_dec['销售数量'] > 0].groupby('订单号').size()
    xhs_dec['销售数量>0的行数'] = xhs_dec['订单号'].map(positive_counts).fillna(0)
    
    def calc_freight_fee(row):
        sales_qty = row['销售数量']
        freight = row['运费_数值']
        
        if sales_qty < 0:
            return freight
        elif sales_qty > 0:
            positive_count = row['销售数量>0的行数']
            if positive_count > 0:
                return freight / positive_count
            else:
                return 0
        else:
            return 0
    
    xhs_dec['收：价外收费'] = xhs_dec.apply(calc_freight_fee, axis=1)
    
    # 扣：平台佣金用（取负值）
    xhs_dec['扣：平台佣金用'] = -xhs_dec['佣金总额'].apply(to_float)
    
    # 扣：分销佣金（取负值）
    xhs_dec['扣：分销佣金'] = -xhs_dec['分销佣金'].apply(to_float)
    
    # 扣其它费用
    xhs_dec['扣其它费用'] = 0
    
    # 应到账金额
    xhs_dec['应到账金额'] = (
        xhs_dec['应收客户'] + 
        xhs_dec['应收平台'] + 
        xhs_dec['收：价外收费'] - 
        xhs_dec['扣：平台佣金用'] - 
        xhs_dec['扣：分销佣金'] - 
        xhs_dec['扣其它费用']
    )
    
    # ============================================================
    # 创建Excel文件（带公式）
    # ============================================================
    wb = Workbook()
    ws = wb.active
    ws.title = '小红书-结算账单'
    
    # 定义计算字段列（A-O列）
    calc_columns = ['年', '月', '订单号', '订单行数', '订单序位', '平台商品编码', '商品编码', 
                    '销售数量', '应收客户', '应收平台', '收：价外收费', '扣：平台佣金用', 
                    '扣：分销佣金', '扣其它费用', '应到账金额']
    
    # 原始数据列（Q-BD列，P列为空）
    raw_data_columns = [
        '订单号', '售后单号', '下单时间', '完成时间', '结算时间', '交易类型', '结算账户', '动账金额',
        '商品名称', '类目', 'SKU条码', '规格ID', '商品数量', '计佣基数', '商品实付/实退', '优惠类型',
        '商家优惠', '平台优惠补贴', '平台运费补贴', '佣金率', '返利率', '佣金总额', '计税价格(含税)',
        '计税价格(未税)', '税率', '跨境税代缴', '商品税金', '卖家CPS佣金率', '分销佣金',
        '推广达人ID', '达人昵称', '带货类型', '代运营服务商佣金', '代开发服务商佣金',
        '运费', '运费税金', '支付渠道费', '花呗分期手续费', '国补订单毛保金额', '备注'
    ]
    
    # 写入表头（第2行）
    for col_idx, col_name in enumerate(calc_columns, 1):
        ws.cell(row=2, column=col_idx, value=col_name)
    
    # P列（第16列）为空
    ws.cell(row=2, column=16, value='')
    
    # 原始数据表头（从Q列=第17列开始）
    for col_idx, col_name in enumerate(raw_data_columns, 17):
        ws.cell(row=2, column=col_idx, value=col_name)
    
    # 写入数据（从第3行开始）
    for row_idx, (_, row) in enumerate(xhs_dec.iterrows(), 3):
        # A列：年
        ws.cell(row=row_idx, column=1, value=year)
        
        # B列：月
        ws.cell(row=row_idx, column=2, value=month)
        
        # C列：订单号（引用Q列）
        ws.cell(row=row_idx, column=3, value=f'=Q{row_idx}')
        
        # D列：订单行数（公式：COUNTIF(Q:Q,Q{row})）
        ws.cell(row=row_idx, column=4, value=f'=COUNTIF(Q:Q,Q{row_idx})')
        
        # E列：订单序位（公式：COUNTIF($Q$3:Q{row},Q{row})）
        ws.cell(row=row_idx, column=5, value=f'=COUNTIF($Q$3:Q{row_idx},Q{row_idx})')
        
        # F列：平台商品编码（值）
        ws.cell(row=row_idx, column=6, value=row['平台商品编码'])
        
        # G列：商品编码（公式）
        ws.cell(row=row_idx, column=7, value=f'=IFERROR(LEFT(F{row_idx},FIND("-",F{row_idx})-1),F{row_idx})')
        
        # H列：销售数量（值）
        ws.cell(row=row_idx, column=8, value=row['销售数量'])
        
        # I列：应收客户（公式：AE{row}）
        ws.cell(row=row_idx, column=9, value=f'=AE{row_idx}')
        
        # J列：应收平台（公式：AH{row}）
        ws.cell(row=row_idx, column=10, value=f'=AH{row_idx}')
        
        # K列：收：价外收费（值）
        ws.cell(row=row_idx, column=11, value=row['收：价外收费'])
        
        # L列：扣：平台佣金用（公式：-AL{row}）
        ws.cell(row=row_idx, column=12, value=f'=-AL{row_idx}')
        
        # M列：扣：分销佣金（公式：-AS{row}）
        ws.cell(row=row_idx, column=13, value=f'=-AS{row_idx}')
        
        # N列：扣其它费用
        ws.cell(row=row_idx, column=14, value=0)
        
        # O列：应到账金额（公式）
        ws.cell(row=row_idx, column=15, value=f'=I{row_idx}+J{row_idx}+K{row_idx}-L{row_idx}-M{row_idx}-N{row_idx}')
        
        # P列：空
        ws.cell(row=row_idx, column=16, value='')
        
        # 原始数据列（从Q列开始）
        for col_offset, col_name in enumerate(raw_data_columns):
            col_idx = 17 + col_offset
            if col_name in xhs_settlement.columns:
                value = row.get(col_name, '')
                if isinstance(value, str) and value.startswith('='):
                    value = "'" + value
                ws.cell(row=row_idx, column=col_idx, value=value)
            elif col_name == '推广达人ID':
                for c in xhs_settlement.columns:
                    if '达人ID' in c:
                        value = row.get(c, '')
                        ws.cell(row=row_idx, column=col_idx, value=value)
                        break
                else:
                    ws.cell(row=row_idx, column=col_idx, value='')
            else:
                ws.cell(row=row_idx, column=col_idx, value='')
    
    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 统计信息
    stats = {
        '总记录数': len(xhs_dec),
        '订单数': xhs_dec['订单号'].nunique(),
        '销售数量合计': xhs_dec['销售数量'].sum(),
        '应收客户合计': xhs_dec['应收客户'].sum(),
        '应到账金额合计': xhs_dec['应到账金额'].sum()
    }
    
    return output, xhs_dec, stats

# ==================== 抖音数据处理（待实现）====================
def process_douyin_data(file1, file2, year, month):
    """处理抖音数据 - 待实现"""
    raise NotImplementedError("抖音数据处理功能开发中...")

# ==================== 视频号数据处理（待实现）====================
def process_shipinhao_data(file1, file2, year, month):
    """处理视频号数据 - 待实现"""
    raise NotImplementedError("视频号数据处理功能开发中...")

# ==================== 主程序 ====================
def main():
    if not check_password():
        return
    
    # 标题
    st.title("📊 电商数据处理系统")
    st.markdown("---")
    
    # 侧边栏：系统设置
    with st.sidebar:
        st.header("⚙️ 系统设置")
        
        st.subheader("支持的平台")
        for platform, config in PLATFORM_CONFIG.items():
            status_text = f"{config['icon']} **{platform}** - {config['status']}"
            if config['enabled']:
                st.success(status_text)
            else:
                st.info(status_text)
        
        st.markdown("---")
        
        # 选择平台
        st.subheader("选择平台")
        enabled_platforms = [p for p, c in PLATFORM_CONFIG.items() if c['enabled']]
        selected_platform = st.selectbox(
            "当前处理平台",
            enabled_platforms,
            help="选择要处理数据的平台"
        )
        
        # 选择月份
        st.subheader("处理月份")
        year = st.number_input("年份", min_value=2020, max_value=2030, value=2025)
        month = st.number_input("月份", min_value=1, max_value=12, value=12)
        
        st.markdown("---")
        
        # 退出登录
        if st.button("🚪 退出登录", use_container_width=True):
            st.session_state.authenticated = False
            st.rerun()
    
    # 主界面
    st.header(f"🚀 步骤1：上传数据文件")
    st.markdown(f"当前平台：**{PLATFORM_CONFIG[selected_platform]['icon']} {selected_platform}**")
    
    uploaded_files = st.file_uploader(
        "请上传2个Excel文件（结算明细 + 订单数据）",
        type=['xlsx', 'xls'],
        accept_multiple_files=True,
        help="系统会自动识别文件类型"
    )
    
    if uploaded_files and len(uploaded_files) >= 2:
        # 识别文件
        files = identify_files(uploaded_files)
        
        if files['settlement'] and files['orders']:
            st.success(f"✅ 已上传 {len(uploaded_files)} 个文件")
            
            col1, col2 = st.columns(2)
            with col1:
                st.info(f"📄 **结算明细**: {files['settlement_name']}")
            with col2:
                st.info(f"📄 **订单数据**: {files['orders_name']}")
            
            # 处理数据
            st.markdown("---")
            st.header("🚀 步骤2：开始处理")
            
            if st.button("开始处理数据", type="primary", use_container_width=True):
                with st.spinner("⏳ 正在读取数据..."):
                    try:
                        processor_name = PLATFORM_CONFIG[selected_platform]['processor']
                        
                        if processor_name == 'process_xiaohongshu':
                            output, result_df, stats = process_xiaohongshu_data(
                                files['settlement'],
                                files['orders'],
                                year,
                                month
                            )
                        elif processor_name == 'process_douyin':
                            output, result_df, stats = process_douyin_data(
                                files['settlement'],
                                files['orders'],
                                year,
                                month
                            )
                        elif processor_name == 'process_shipinhao':
                            output, result_df, stats = process_shipinhao_data(
                                files['settlement'],
                                files['orders'],
                                year,
                                month
                            )
                        
                        st.success("✅ 处理完成！")
                        
                        # 显示统计信息
                        st.subheader("📊 数据统计")
                        cols = st.columns(len(stats))
                        for col, (key, value) in zip(cols, stats.items()):
                            with col:
                                if isinstance(value, (int, float)):
                                    st.metric(key, f"{value:,.2f}" if isinstance(value, float) else f"{value:,}")
                                else:
                                    st.metric(key, value)
                        
                        # 下载按钮
                        st.markdown("---")
                        st.subheader("📥 下载结果")
                        
                        filename = f"{selected_platform}_{year}年{month}月结算账单.xlsx"
                        st.download_button(
                            label="📥 下载Excel文件",
                            data=output,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                        
                    except Exception as e:
                        st.error(f"❌ 处理失败: {str(e)}")
                        with st.expander("查看详细错误信息"):
                            st.code(str(e))
        else:
            st.warning("⚠️ 无法识别文件类型，请确保上传了正确的结算明细和订单数据文件")
    elif uploaded_files:
        st.warning(f"⚠️ 请上传至少2个文件（当前已上传 {len(uploaded_files)} 个）")

if __name__ == "__main__":
    main()
